import os
from pathlib import Path
from typing import Dict, Any, List, Tuple
import openpyxl
import win32com.client
import pythoncom
import psutil

# File paths
INPUT_FILE = Path(r"C:\Users\MOMI9362\Input.xlsx")
DESIGN_FILE = Path(r"C:\Users\MOMI9362\Design.xlsx")
OUTPUT_FILE = Path(r"C:\Users\MOMI9362\Output.xlsx")
DESIGN_CASES_DIR = Path(r"C:\Users\MOMI9362\Design_Cases")

# Parameter mapping for Design.xlsx
PARAM_MAP: Dict[str, Tuple[str, str]] = {
    "width": ("Design", "D26"),
    "length": ("Design", "D27"),
    "cohesion": ("Design", "D20"),
    "phi": ("Design", "D19"),
    "gwt_depth": ("Design", "D33"),
    "burial_depth": ("Design", "D28"),
}

SBC_CELL = ("Design", "B68")

def load_parameter_sets(input_file: Path) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    print("Loaded parameter sets:", rows)
    return rows

def create_output_template(rows: List[Dict[str, Any]]) -> openpyxl.Workbook:
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Results"
    headers = list(rows[0].keys()) + ["Safe_Bearing_Capacity_kN_m2"]
    for col_idx, h in enumerate(headers, 1):
        ws_out.cell(row=1, column=col_idx, value=h)
    return wb_out

def kill_excel_processes():
    for proc in psutil.process_iter(['name']):
        if proc.info['name'].lower() == 'excel.exe':
            proc.kill()
            print("Terminated lingering Excel process")

def process_design_file(param_set: Dict[str, Any], design_file: Path, case_idx: int) -> tuple[float, Path]:
    pythoncom.CoInitialize()
    excel = None
    wb = None
    output_path = DESIGN_CASES_DIR / f"Design_Case_{case_idx:03d}.xlsx"
    try:
        print(f"Opening Excel application for param_set: {param_set}")
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        print(f"Opening workbook: {design_file}")
        wb = excel.Workbooks.Open(str(design_file))
        print("Worksheets in Design.xlsx:", [ws.Name for ws in wb.Worksheets])
        ws = wb.Worksheets("Design")

        for name, (sheet, cell) in PARAM_MAP.items():
            print(f"Writing {name}={param_set[name]} to {cell}")
            ws.Range(cell).Value = param_set[name]

        print("Calculating formulas")
        excel.Calculate()
        sbc = ws.Range(SBC_CELL[1]).Value
        print(f"Read SBC: {sbc}")

        print(f"Saving design file to: {output_path}")
        wb.SaveAs(str(output_path))
        return sbc, output_path
    except Exception as e:
        print(f"Error processing design file for case {case_idx}: {e}")
        raise
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
                print("Closed workbook")
            if excel is not None:
                excel.Quit()
                print("Quit Excel application")
        except Exception as e:
            print(f"Error during cleanup: {e}")
        pythoncom.CoUninitialize()
        kill_excel_processes()

def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")
    if not DESIGN_FILE.exists():
        raise FileNotFoundError(f"Design file not found: {DESIGN_FILE}")

    # Create output directory for design cases
    DESIGN_CASES_DIR.mkdir(exist_ok=True)
    print(f"Created/using design cases directory: {DESIGN_CASES_DIR}")

    rows = load_parameter_sets(INPUT_FILE)
    if not rows:
        raise ValueError("No parameter sets loaded from Input.xlsx")

    valid_rows = [row for row in rows if all(row[key] is not None for key in ['width', 'length', 'cohesion', 'phi'])]
    print(f"Valid parameter sets: {len(valid_rows)} out of {len(rows)}")

    wb_out = create_output_template(valid_rows)
    ws_out = wb_out.active

    for r_idx, param_set in enumerate(valid_rows, start=2):
        print(f"Processing row {r_idx}: {param_set}")
        try:
            sbc, saved_design_path = process_design_file(param_set, DESIGN_FILE, r_idx - 1)
            print(f"Row {r_idx}: SBC = {sbc}, Saved design file: {saved_design_path}")
        except Exception as e:
            print(f"Failed to process row {r_idx}: {e}")
            continue

        for c_idx, key in enumerate(param_set.keys(), start=1):
            ws_out.cell(row=r_idx, column=c_idx, value=param_set[key])
            print(f"Writing to Output.xlsx: row={r_idx}, column={c_idx}, value={param_set[key]}")
        ws_out.cell(row=r_idx, column=len(param_set)+1, value=sbc)
        print(f"Writing SBC to Output.xlsx: row={r_idx}, column={len(param_set)+1}, value={sbc}")

    wb_out.save(OUTPUT_FILE)
    print(f"Done. Results written to {OUTPUT_FILE.resolve()}")
    print(f"Design files saved to {DESIGN_CASES_DIR.resolve()}")

if __name__ == "__main__":
    main()